import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 체크리스트 데이터 예시 - 반영여부와 오시공여부 포함
rows = [
    ["구분", "점검항목(설계지침)", "세부기준/치수", "반영여부", "오시공여부", "현장적용여부(체크)", "비고(특이사항/사진첨부)"],
    ["현관/복도", "현관 폭", "1,000mm 이상(84㎡ 미만), 1,200mm 이상(84㎡ 이상)", "□", "□", "□", ""],
    ["현관/복도", "신발장 깊이", "400mm 이상(일반), 600mm 이상(깊은 수납장)", "□", "□", "□", ""],
    ["현관/복도", "현관문 규격", "1,100×2,100mm(방화문)", "□", "□", "□", ""],
    ["현관/복도", "현관문틀 단열", "문틀 주변 cold bridge 방지, 단열재 연속시공", "□", "□", "□", ""],
    ["거실", "거실 전면폭", "3,000mm 이상(60㎡ 미만), 3,600mm 이상(80㎡ 이상)", "□", "□", "□", ""],
    ["거실", "창호 규격", "창호 전체길이, 분합문 조합 등", "□", "□", "□", ""],
    ["거실", "창호 주변 단열", "외부창호 상하부턱 단열조치", "□", "□", "□", ""],
    ["주방/식당", "주방가구 길이", "3,000mm 이상(60㎡ 미만), 3,600mm 이상(85㎡ 이상)", "□", "□", "□", ""],
    ["주방/식당", "식탁 설치 공간", "4인용(1,400~1,600mm), 6인용(1,800~2,000mm)", "□", "□", "□", ""],
    ["주방/식당", "개수대 설치장 깊이", "700mm 확보", "□", "□", "□", ""],
    ["주방/식당", "주방창 규격", "1,200×1,100mm 이상(상부장 미설치)", "□", "□", "□", ""],
    ["욕실", "욕실 슬라브 두께", "180mm(거실보다 30mm 낮춤)", "□", "□", "□", ""],
    ["욕실", "욕실 창호", "700×2,100mm", "□", "□", "□", ""],
    ["욕실", "욕실 바닥 레벨", "거실과 60mm 단차", "□", "□", "□", ""],
    ["욕실", "욕실 방수", "액체방수 적용", "□", "□", "□", ""],
    ["발코니", "세탁기 위치", "폭 900mm, 깊이 900mm 이상", "□", "□", "□", ""],
    ["발코니", "실외기실 크기", "1,450×1,150mm 이상", "□", "□", "□", ""],
    ["발코니", "난간 높이", "FL+1,200mm", "□", "□", "□", ""],
    ["발코니", "발코니 수전 설치", "내측 벽에 부착(동파방지)", "□", "□", "□", ""],
    ["공용공간", "장애인 램프 폭", "1,500mm 이상", "□", "□", "□", ""],
    ["공용공간", "계단실 유효폭", "2,600mm 이상", "□", "□", "□", ""],
    ["공용공간", "E/V홀 유효폭", "1,800mm 이상", "□", "□", "□", ""],
    ["공용공간", "우편물 수취함", "최상단 1,800mm 이하, 최하단 400mm 이상", "□", "□", "□", ""],
    ["기타", "단열/결로방지", "현관문틀, 창호주변, 바닥 등 단열재 연속시공", "□", "□", "□", ""],
    ["기타", "층간소음완충재", "분합문틀 하부 취부", "□", "□", "□", ""],
    ["기타", "방수 시공", "지붕, 옥탑, 욕실, 발코니 방수상세", "□", "□", "□", ""],
]

def create_checklist_excel(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "체크리스트"
    
    # 데이터 입력
    for row in rows:
        ws.append(row)
    
    # 스타일 적용
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 스타일 적용
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 열 너비 조정
    column_widths = [15, 25, 35, 12, 12, 20, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 모든 셀에 테두리 적용
    for row in ws.iter_rows(min_row=1, max_row=len(rows)):
        for cell in row:
            cell.border = thin_border
    
    wb.save(filename)

if __name__ == "__main__":
    create_checklist_excel("체크리스트_상세.xlsx") 